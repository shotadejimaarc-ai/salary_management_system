import os
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib import pagesizes
from reportlab.lib.units import mm
from openpyxl.styles import Border, Side


# ===============================
# 共通：値取得（0問題対策済）
# ===============================
def _get_salary_values(salary_result):

    staff_type = salary_result.get("type", "")
    personal_sales_amount = salary_result.get("personal_sales_amount", 0)
    personal_sales_f = salary_result.get("personal_sales_f", 0)

    # ★ ここ修正
    organization_sales_amount = salary_result.get("org_sales_amount", 0)
    organization_sales_f = salary_result.get("org_sales_f", 0)

    # ★ ここ修正
    commission_rate = salary_result.get("commission_rate", 0)

    # ★ ここ修正
    total_salary = salary_result.get("total", 0)

    staff_type_label = "スタッフ" if staff_type == "staff" else "バイト"

    return {
        "staff_type_label": staff_type_label,
        "personal_sales_amount": personal_sales_amount,
        "personal_sales_f": personal_sales_f,
        "organization_sales_amount": organization_sales_amount,
        "organization_sales_f": organization_sales_f,
        "organization_rate": commission_rate * 100,  # %表示用
        "total_salary": total_salary,
    }


# ===============================
# Excel 出力（プロ仕様）
# ===============================
def export_salary_excel(staff_name, salary_result, detail_df, output_path):

    # ===== list対策 =====
    if isinstance(detail_df, list):
        detail_df = pd.DataFrame(detail_df)
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    values = _get_salary_values(salary_result)

    wb = Workbook()
    ws = wb.active
    ws.title = "給与明細"

    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    strong_font = Font(bold=True, size=14)
    strong_fill = PatternFill(start_color="FFF2CC", fill_type="solid")

    summary_data = [
        ["担当者区分", values["staff_type_label"]],
        ["個人売上金額", f"¥{values['personal_sales_amount']:,}"],
        ["個人売上F", f"{values['personal_sales_f']:,}F"],
        ["組織売上金額", f"¥{values['organization_sales_amount']:,}"],
        ["組織売上F", f"{values['organization_sales_f']:,}F"],
        ["適用レート", f"{values['organization_rate']}%"],
        ["総支給額", f"¥{values['total_salary']:,}"],
    ]

    row = 1
    for label, value in summary_data:

    # ラベル（A-B結合）
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=label)

        # 値（C-D結合）
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3, value=value)

        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).font = header_font

        # 中央寄せにしたいなら追加
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        if label == "総支給額":

            # フォント大きめに
            big_font = Font(bold=True, size=18)

            ws.cell(row=row, column=1).font = big_font
            ws.cell(row=row, column=3).font = big_font

            ws.cell(row=row, column=1).fill = strong_fill
            ws.cell(row=row, column=3).fill = strong_fill

            # 横中央 + 縦中央
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            ws.cell(row=row, column=3).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            # 太枠
            ws.cell(row=row, column=1).border = thick_border
            ws.cell(row=row, column=2).border = thick_border
            ws.cell(row=row, column=3).border = thick_border
            ws.cell(row=row, column=4).border = thick_border

        row += 1

    row += 2

    # ===== 列幅調整 =====
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16

    # ===== 行の高さ調整 =====
    ws.row_dimensions[7].height = 28
    ws.row_dimensions[10].height = 28

    headers = ["営業日", "担当者名", "カテゴリ", "商品名", "売上金額", "親分配率", "分配後金額"]

    detail_header_row = row

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row += 1

    # ===== 明細 =====
    if not detail_df.empty:
        for _, r in detail_df.iterrows():
            ws.append([
                r.get("date", ""),
                r.get("staff_name", ""),
                r.get("category", ""),
                r.get("product_name", ""),
                r.get("sales_amount", 0),
                r.get("rate", 0),
                r.get("calculated_amount", 0),
            ])
    
    
    # ===== 明細に罫線 =====
    if ws.max_row >= detail_header_row:

        for r in ws.iter_rows(
            min_row=detail_header_row,
            max_row=ws.max_row,
            min_col=1,
            max_col=7
        ):
            for cell in r:
                cell.border = thin_border

    # ===== BytesIO対応 =====
    if hasattr(output_path, "write"):
        wb.save(output_path)
    else:
        wb.save(str(output_path))


# ===============================
# PDF 出力（本物会社給与明細）
# ===============================
def export_salary_pdf(staff_name, salary_result, detail_df, output_path):

    if isinstance(detail_df, list):
        detail_df = pd.DataFrame(detail_df)
    values = _get_salary_values(salary_result)

    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))

    doc = SimpleDocTemplate(output_path, pagesize=pagesizes.A4)
    elements = []

    title_style = ParagraphStyle(
        name="Title",
        fontName="HeiseiKakuGo-W5",
        fontSize=16,
        alignment=1
    )

    elements.append(Paragraph("給与明細書", title_style))
    elements.append(Spacer(1, 15))

    # ===== 基本情報 =====
    summary_table_data = [
        ["氏名", staff_name],
        ["担当者区分", values["staff_type_label"]],
        ["個人売上金額", f"¥{values['personal_sales_amount']:,}"],
        ["個人売上F", f"{values['personal_sales_f']:,}F"],
        ["組織売上金額", f"¥{values['organization_sales_amount']:,}"],
        ["組織売上F", f"{values['organization_sales_f']:,}F"],
        ["適用レート", f"{values['organization_rate']}%"],
        ["総支給額", f"¥{values['total_salary']:,}"],
    ]

    summary_table = Table(summary_table_data, colWidths=[50*mm, 60*mm])

    summary_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,7), (-1,7), colors.yellow),
        ("FONTNAME", (0,0), (-1,-1), "HeiseiKakuGo-W5"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # ===== 明細 =====
    detail_data = [["営業日", "カテゴリ", "商品名", "売上金額", "親分配率", "F計上額"]]

    for _, r in detail_df.iterrows():
        detail_data.append([
            r.get("date", ""),
            r.get("category", ""),
            r.get("product_name", ""),
            f"{r.get('sales_amount',0):,}",
            f"{r.get('rate',0)}%",
            f"{r.get('calculated_amount',0):,}",
        ])

    detail_table = Table(detail_data, repeatRows=1)

    detail_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.lightblue),
        ("FONTNAME", (0,0), (-1,-1), "HeiseiKakuGo-W5"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
    ]))

    elements.append(detail_table)

    doc.build(elements)

def generate_salary_excel_pdf_full(
    staff_name,
    salary_result,
    month_str,
    detail_df
):

    # ======================
    # Excelをメモリ上で生成
    # ======================
    excel_buffer = io.BytesIO()
    export_salary_excel(staff_name, salary_result, detail_df, excel_buffer)
    excel_bytes = excel_buffer.getvalue()

    # ======================
    # PDFをメモリ上で生成
    # ======================
    pdf_buffer = io.BytesIO()
    export_salary_pdf(staff_name, salary_result, detail_df, pdf_buffer)
    pdf_bytes = pdf_buffer.getvalue()

    return excel_bytes, pdf_bytes